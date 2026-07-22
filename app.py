import customtkinter as ctk
from tkcalendar import DateEntry, Calendar
from datetime import timedelta, datetime
from PIL import Image, ImageDraw
import json
import os
from pathlib import Path
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
import calendar

# ==========================================
# CONFIGURAÇÃO
# ==========================================

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

ARQUIVO_HISTORICO = "historico_consultas.json"
ARQUIVO_CONFIGURACAO = "configuracao.json"

# Configuração padrão
CONFIG_PADRAO = {
    "data_referencia": None,
    "period_start": "quinta",  # quinta, sexta, sábado
    "period_end": "domingo",   # domingo, segunda
    "tema": "light",
    "feriados": []
}

# ==========================================
# GERENCIADOR DE DADOS
# ==========================================

class GerenciadorDados:
    def __init__(self):
        self.config = self.carregar_config()
        self.historico = self.carregar_historico()
    
    def carregar_config(self):
        if os.path.exists(ARQUIVO_CONFIGURACAO):
            try:
                with open(ARQUIVO_CONFIGURACAO, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return CONFIG_PADRAO.copy()
        return CONFIG_PADRAO.copy()
    
    def salvar_config(self):
        with open(ARQUIVO_CONFIGURACAO, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2, default=str)
    
    def carregar_historico(self):
        if os.path.exists(ARQUIVO_HISTORICO):
            try:
                with open(ARQUIVO_HISTORICO, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return []
        return []
    
    def salvar_historico(self):
        with open(ARQUIVO_HISTORICO, 'w', encoding='utf-8') as f:
            json.dump(self.historico, f, ensure_ascii=False, indent=2, default=str)
    
    def adicionar_consulta(self, data):
        self.historico.append({
            "data": data.strftime('%d/%m/%Y'),
            "timestamp": datetime.now().isoformat()
        })
        self.salvar_historico()

dados = GerenciadorDados()

# ==========================================
# FUNÇÕES DE CÁLCULO
# ==========================================

def obter_sabado(data):
    """Obtém o sábado da semana atual"""
    wd = data.weekday()
    
    if wd == 5:  # sábado
        return data
    
    if wd == 6:  # domingo
        return data - timedelta(days=1)
    
    return data - timedelta(days=wd + 2)


def calcular_periodos(data_inicio, num_semanas=52):
    """Calcula períodos baseado na data de início"""
    sabado = obter_sabado(data_inicio)
    periodos = []
    
    for i in range(num_semanas):
        inicio = sabado + timedelta(days=i * 14)
        fim = inicio + timedelta(days=1)
        tem_filhas = (i % 2 == 0)
        
        periodos.append({
            'inicio': inicio,
            'fim': fim,
            'tem_filhas': tem_filhas,
            'semana': i
        })
    
    return periodos


def obter_proximos_periodos(data_referencia, num_periodos=5):
    """Obtém os próximos períodos com as filhas"""
    periodos = calcular_periodos(data_referencia, num_semanas=52)
    proximos = [p for p in periodos if p['tem_filhas'] and p['inicio'] >= data_referencia]
    return proximos[:num_periodos]


def calcular_dias_faltantes(data_inicio):
    """Calcula quantos dias faltam até a data"""
    hoje = datetime.now().date()
    if data_inicio < hoje:
        return 0
    diferenca = (data_inicio - hoje).days
    return diferenca


def verificar_evento(data_evento, data_referencia):
    """Verifica se haverá guarda no dia do evento"""
    referencia = data_referencia
    evento = data_evento
    
    sab_ref = obter_sabado(referencia)
    sab_evento = obter_sabado(evento)
    
    diferenca_semanas = (sab_evento - sab_ref).days // 7
    
    tem_filhas = (diferenca_semanas % 2 == 0)
    
    return {
        'tem_filhas': tem_filhas,
        'inicio': sab_evento,
        'fim': sab_evento + timedelta(days=1)
    }


def contar_periodos_ano(data_referencia, ano, tem_filhas=True):
    """Conta quantos períodos com/sem filhas em um determinado ano"""
    periodos = calcular_periodos(data_referencia, num_semanas=260)  # ~5 anos
    
    contagem = 0
    for p in periodos:
        if p['inicio'].year == ano and p['tem_filhas'] == tem_filhas:
            contagem += 1
    
    return contagem


# ==========================================
# INTERFACE - DASHBOARD
# ==========================================

def criar_dashboard():
    """Cria a tela inicial do dashboard"""
    janela = ctk.CTkToplevel(app)
    janela.title("Dashboard")
    janela.geometry("700x600")
    janela.resizable(False, False)
    
    # Título
    titulo = ctk.CTkLabel(
        janela,
        text="👧 CONTROLE DE GUARDA",
        font=("Segoe UI", 24, "bold")
    )
    titulo.pack(pady=20)
    
    try:
        referencia = data_referencia.get_date()
        dados.adicionar_consulta(referencia)
        
        # Próximo período
        proximos = obter_proximos_periodos(referencia, 1)
        
        if proximos:
            proximo = proximos[0]
            dias_faltantes = calcular_dias_faltantes(proximo['inicio'].date() if isinstance(proximo['inicio'], datetime) else proximo['inicio'])
            
            card_proximo = ctk.CTkFrame(janela, fg_color="#2E7D32", corner_radius=10)
            card_proximo.pack(padx=20, pady=10, fill="x")
            
            ctk.CTkLabel(
                card_proximo,
                text="📅 Próximo Período Com Filhas",
                font=("Segoe UI", 12, "bold"),
                text_color="white"
            ).pack(pady=(10, 0))
            
            ctk.CTkLabel(
                card_proximo,
                text=f"{proximo['inicio'].strftime('%d/%m/%Y')} até {proximo['fim'].strftime('%d/%m/%Y')}",
                font=("Segoe UI", 14, "bold"),
                text_color="white"
            ).pack()
            
            ctk.CTkLabel(
                card_proximo,
                text=f"⏳ Faltam {dias_faltantes} dias",
                font=("Segoe UI", 11),
                text_color="white"
            ).pack(pady=(0, 10))
        
        # Total em 2026
        ano_atual = datetime.now().year
        total_com = contar_periodos_ano(referencia, ano_atual, tem_filhas=True)
        total_sem = contar_periodos_ano(referencia, ano_atual, tem_filhas=False)
        
        card_total = ctk.CTkFrame(janela, fg_color="#1F6AA5", corner_radius=10)
        card_total.pack(padx=20, pady=10, fill="x")
        
        ctk.CTkLabel(
            card_total,
            text="✅ TOTAL EM " + str(ano_atual),
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).pack(pady=(10, 5))
        
        frame_contagem = ctk.CTkFrame(card_total, fg_color="#1F6AA5")
        frame_contagem.pack(fill="x", padx=10, pady=(0, 10))
        
        ctk.CTkLabel(
            frame_contagem,
            text=f"🟢 Com filhas: {total_com}",
            font=("Segoe UI", 11),
            text_color="white"
        ).pack(side="left", padx=15)
        
        ctk.CTkLabel(
            frame_contagem,
            text=f"🔴 Sem filhas: {total_sem}",
            font=("Segoe UI", 11),
            text_color="white"
        ).pack(side="left", padx=15)
        
        # Próximos 5 períodos
        card_proximos = ctk.CTkFrame(janela, fg_color="#F5F5F5", corner_radius=10)
        card_proximos.pack(padx=20, pady=10, fill="both", expand=True)
        
        ctk.CTkLabel(
            card_proximos,
            text="📅 Próximos 5 Períodos",
            font=("Segoe UI", 12, "bold")
        ).pack(pady=(10, 5))
        
        # Lista de períodos
        proximos_5 = obter_proximos_periodos(referencia, 5)
        
        for i, periodo in enumerate(proximos_5, 1):
            dias_faltam = calcular_dias_faltantes(periodo['inicio'].date() if isinstance(periodo['inicio'], datetime) else periodo['inicio'])
            
            frame_item = ctk.CTkFrame(card_proximos, fg_color="white", corner_radius=8)
            frame_item.pack(padx=10, pady=5, fill="x")
            
            ctk.CTkLabel(
                frame_item,
                text=f"{i}. {periodo['inicio'].strftime('%d/%m/%Y')} → {periodo['fim'].strftime('%d/%m/%Y')} (Faltam {dias_faltam} dias)",
                font=("Segoe UI", 10),
                text_color="black"
            ).pack(pady=5, anchor="w", padx=10)
    
    except Exception as e:
        ctk.CTkLabel(
            janela,
            text=f"Erro ao carregar dashboard: {str(e)}",
            text_color="red"
        ).pack(pady=20)


# ==========================================
# INTERFACE - CALENDÁRIO VISUAL
# ==========================================

def criar_calendario_visual():
    """Cria visualização de calendário colorido"""
    janela = ctk.CTkToplevel(app)
    janela.title("Calendário Visual")
    janela.geometry("900x700")
    
    try:
        referencia = data_referencia.get_date()
        
        # Frame para seleção de mês/ano
        frame_nav = ctk.CTkFrame(janela)
        frame_nav.pack(pady=10, fill="x", padx=10)
        
        mes_atual = datetime.now().month
        ano_atual = datetime.now().year
        
        mes_var = ctk.StringVar(value=str(mes_atual))
        ano_var = ctk.StringVar(value=str(ano_atual))
        
        ctk.CTkLabel(frame_nav, text="Mês:").pack(side="left", padx=5)
        
        combo_mes = ctk.CTkComboBox(
            frame_nav,
            values=[str(i) for i in range(1, 13)],
            variable=mes_var,
            width=50
        )
        combo_mes.pack(side="left", padx=5)
        
        ctk.CTkLabel(frame_nav, text="Ano:").pack(side="left", padx=5)
        
        combo_ano = ctk.CTkComboBox(
            frame_nav,
            values=[str(i) for i in range(ano_atual, ano_atual + 6)],
            variable=ano_var,
            width=80
        )
        combo_ano.pack(side="left", padx=5)
        
        # Frame para exibir calendário
        frame_calendario = ctk.CTkScrollableFrame(janela)
        frame_calendario.pack(padx=20, pady=10, fill="both", expand=True)
        
        def atualizar_calendario():
            # Limpar frame anterior
            for widget in frame_calendario.winfo_children():
                widget.destroy()
            
            mes = int(mes_var.get())
            ano = int(ano_var.get())
            
            # Título
            nome_mes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                       "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
            
            ctk.CTkLabel(
                frame_calendario,
                text=f"{nome_mes[mes-1].upper()} {ano}",
                font=("Segoe UI", 18, "bold")
            ).pack(pady=15)
            
            # Cabeçalho da semana
            dias_semana = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
            frame_header = ctk.CTkFrame(frame_calendario)
            frame_header.pack()
            
            for dia in dias_semana:
                ctk.CTkLabel(
                    frame_header,
                    text=dia,
                    font=("Segoe UI", 10, "bold"),
                    width=50
                ).pack(side="left", padx=5)
            
            # Dias do mês
            cal = calendar.monthcalendar(ano, mes)
            periodos = calcular_periodos(referencia, num_semanas=260)
            
            for semana in cal:
                frame_week = ctk.CTkFrame(frame_calendario)
                frame_week.pack(pady=5)
                
                for dia in semana:
                    if dia == 0:
                        ctk.CTkLabel(
                            frame_week,
                            text="",
                            width=50
                        ).pack(side="left", padx=5)
                    else:
                        data_dia = datetime(ano, mes, dia).date()
                        
                        # Verificar se tem filhas nesse dia
                        tem_filhas = False
                        for periodo in periodos:
                            inicio = periodo['inicio'].date() if isinstance(periodo['inicio'], datetime) else periodo['inicio']
                            fim = periodo['fim'].date() if isinstance(periodo['fim'], datetime) else periodo['fim']
                            
                            if inicio <= data_dia <= fim and periodo['tem_filhas']:
                                tem_filhas = True
                                break
                        
                        cor = "#2E7D32" if tem_filhas else "#C62828"
                        
                        ctk.CTkLabel(
                            frame_week,
                            text=str(dia).zfill(2),
                            font=("Segoe UI", 10, "bold"),
                            fg_color=cor,
                            text_color="white",
                            width=50,
                            corner_radius=5
                        ).pack(side="left", padx=5)
        
        # Botão para atualizar
        ctk.CTkButton(
            frame_nav,
            text="Atualizar",
            command=atualizar_calendario,
            width=100
        ).pack(side="right", padx=5)
        
        atualizar_calendario()
    
    except Exception as e:
        ctk.CTkLabel(
            janela,
            text=f"Erro: {str(e)}",
            text_color="red"
        ).pack(pady=20)


# ==========================================
# INTERFACE - TABELA DE PERÍODOS
# ==========================================

def listar_periodos(num_semanas=52):
    """Lista períodos em formato de tabela"""
    janela = ctk.CTkToplevel(app)
    janela.title("Períodos de Guarda")
    janela.geometry("900x600")
    
    try:
        referencia = data_referencia.get_date()
        
        # Frame com seleção de período
        frame_opcoes = ctk.CTkFrame(janela)
        frame_opcoes.pack(pady=10, fill="x", padx=10)
        
        ctk.CTkLabel(frame_opcoes, text="Mostrar:", font=("Segoe UI", 12)).pack(side="left", padx=5)
        
        def atualizar(semanas):
            listar_periodos(semanas)
            janela.destroy()
        
        ctk.CTkButton(
            frame_opcoes,
            text="1 Ano (52 semanas)",
            command=lambda: atualizar(52),
            width=150
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            frame_opcoes,
            text="2 Anos",
            command=lambda: atualizar(104),
            width=150
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            frame_opcoes,
            text="5 Anos",
            command=lambda: atualizar(260),
            width=150
        ).pack(side="left", padx=5)
        
        # Título
        titulo = ctk.CTkLabel(
            janela,
            text="PERÍODOS DE GUARDA",
            font=("Segoe UI", 18, "bold")
        )
        titulo.pack(pady=10)
        
        # Textbox com os períodos
        texto = ctk.CTkTextbox(
            janela,
            width=850,
            height=500,
            font=("Consolas", 11)
        )
        texto.pack(padx=20, pady=10, fill="both", expand=True)
        
        # Header
        texto.insert("end", f"{'Nº':<4} {'Início':<12} {'Fim':<12} {'Status':<20} {'Dias Faltam':<12}\n")
        texto.insert("end", "-" * 80 + "\n")
        
        periodos = calcular_periodos(referencia, num_semanas)
        
        for i, periodo in enumerate(periodos, 1):
            inicio_str = periodo['inicio'].strftime('%d/%m/%Y')
            fim_str = periodo['fim'].strftime('%d/%m/%Y')
            status = "✅ COM AS FILHAS" if periodo['tem_filhas'] else "❌ SEM AS FILHAS"
            dias_faltam = calcular_dias_faltantes(periodo['inicio'].date() if isinstance(periodo['inicio'], datetime) else periodo['inicio'])
            
            linha = f"{i:<4} {inicio_str:<12} {fim_str:<12} {status:<20} {dias_faltam} dias\n"
            texto.insert("end", linha)
        
        texto.configure(state="disabled")
    
    except Exception as e:
        ctk.CTkLabel(
            janela,
            text=f"Erro: {str(e)}",
            text_color="red"
        ).pack(pady=20)


# ==========================================
# INTERFACE - PESQUISA POR DATA
# ==========================================

def criar_pesquisa_data():
    """Cria janela de pesquisa por data"""
    janela = ctk.CTkToplevel(app)
    janela.title("Pesquisar Data")
    janela.geometry("600x400")
    janela.resizable(False, False)
    
    # Título
    ctk.CTkLabel(
        janela,
        text="🔍 PESQUISAR DATA",
        font=("Segoe UI", 20, "bold")
    ).pack(pady=20)
    
    # Frame principal
    frame = ctk.CTkFrame(janela)
    frame.pack(padx=30, pady=20, fill="both", expand=True)
    
    # Label
    ctk.CTkLabel(
        frame,
        text="Digite uma data:",
        font=("Segoe UI", 14)
    ).pack(pady=(20, 10))
    
    # DateEntry
    data_pesquisa = DateEntry(
        frame,
        locale="pt_BR",
        date_pattern="dd/mm/yyyy",
        width=20,
        background="#1F6AA5",
        foreground="white",
        borderwidth=2
    )
    data_pesquisa.pack(pady=10)
    
    # Label resultado
    resultado_label = ctk.CTkLabel(
        frame,
        text="",
        font=("Segoe UI", 12),
        justify="left",
        fg_color="transparent"
    )
    resultado_label.pack(pady=20)
    
    def pesquisar():
        try:
            referencia = data_referencia.get_date()
            data_evento = data_pesquisa.get_date()
            
            resultado = verificar_evento(data_evento, referencia)
            
            if resultado['tem_filhas']:
                status = "✅ Você estará com suas filhas"
                cor = "#2E7D32"
            else:
                status = "❌ Você NÃO estará com suas filhas"
                cor = "#C62828"
            
            resultado_text = (
                f"{status}\n\n"
                f"Período: {resultado['inicio'].strftime('%d/%m/%Y')} até {resultado['fim'].strftime('%d/%m/%Y')}\n"
                f"Dias faltando: {calcular_dias_faltantes(resultado['inicio'].date() if isinstance(resultado['inicio'], datetime) else resultado['inicio'])} dias"
            )
            
            resultado_label.configure(text=resultado_text, text_color="white", fg_color=cor)
        except Exception as e:
            resultado_label.configure(text=f"Erro: {str(e)}", text_color="red", fg_color="transparent")
    
    # Botão pesquisar
    botao_pesquisar = ctk.CTkButton(
        frame,
        text="🔎 Pesquisar",
        command=pesquisar,
        width=200,
        height=40,
        font=("Segoe UI", 14, "bold")
    )
    botao_pesquisar.pack(pady=20)


# ==========================================
# INTERFACE - EXPORTAR PDF
# ==========================================

def exportar_pdf():
    """Exporta calendário em PDF"""
    try:
        from tkinter import filedialog
        
        referencia = data_referencia.get_date()
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"calendario_guarda_{datetime.now().year}.pdf"
        )
        
        if not arquivo:
            return
        
        # Criar PDF
        doc = SimpleDocTemplate(arquivo, pagesize=letter)
        elements = []
        
        # Título
        styles = getSampleStyleSheet()
        style_titulo = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor='#2E7D32',
            spaceAfter=30,
            alignment=1
        )
        
        elements.append(Paragraph("CALENDÁRIO DE GUARDA DAS FILHAS", style_titulo))
        elements.append(Paragraph(f"Ano: {datetime.now().year}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabela de períodos
        periodos = calcular_periodos(referencia, num_semanas=52)
        
        data_tabela = [["Período", "Início", "Fim", "Status"]]
        
        for i, periodo in enumerate(periodos, 1):
            status = "✅ COM AS FILHAS" if periodo['tem_filhas'] else "❌ SEM AS FILHAS"
            data_tabela.append([
                str(i),
                periodo['inicio'].strftime('%d/%m/%Y'),
                periodo['fim'].strftime('%d/%m/%Y'),
                status
            ])
        
        table = Table(data_tabela)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), '#2E7D32'),
            ('TEXTCOLOR', (0, 0), (-1, 0), 'white'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, 'black'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), ['white', '#f0f0f0'])
        ]))
        
        elements.append(table)
        
        doc.build(elements)
        
        ctk.CTkLabel(
            app,
            text=f"✅ PDF exportado: {arquivo}",
            text_color="green",
            font=("Segoe UI", 10)
        ).pack(pady=5)
    
    except Exception as e:
        ctk.CTkLabel(
            app,
            text=f"❌ Erro ao exportar PDF: {str(e)}",
            text_color="red",
            font=("Segoe UI", 10)
        ).pack(pady=5)


# ==========================================
# INTERFACE - EXPORTAR EXCEL
# ==========================================

def exportar_excel():
    """Exporta calendário em Excel"""
    try:
        from tkinter import filedialog
        
        referencia = data_referencia.get_date()
        arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"calendario_guarda_{datetime.now().year}.xlsx"
        )
        
        if not arquivo:
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Períodos"
        
        # Headers
        headers = ["Período", "Início", "Fim", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        periodos = calcular_periodos(referencia, num_semanas=52)
        
        for row, periodo in enumerate(periodos, 2):
            status = "✅ COM AS FILHAS" if periodo['tem_filhas'] else "❌ SEM AS FILHAS"
            
            ws.cell(row=row, column=1).value = row - 1
            ws.cell(row=row, column=2).value = periodo['inicio'].strftime('%d/%m/%Y')
            ws.cell(row=row, column=3).value = periodo['fim'].strftime('%d/%m/%Y')
            ws.cell(row=row, column=4).value = status
            
            # Colorir linhas
            cor = "C8E6C9" if periodo['tem_filhas'] else "FFCDD2"
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=cor,
                    end_color=cor,
                    fill_type="solid"
                )
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        wb.save(arquivo)
        
        ctk.CTkLabel(
            app,
            text=f"✅ Excel exportado: {arquivo}",
            text_color="green",
            font=("Segoe UI", 10)
        ).pack(pady=5)
    
    except Exception as e:
        ctk.CTkLabel(
            app,
            text=f"❌ Erro ao exportar Excel: {str(e)}",
            text_color="red",
            font=("Segoe UI", 10)
        ).pack(pady=5)


# ==========================================
# INTERFACE - COPIAR PARA WHATSAPP
# ==========================================

def copiar_whatsapp():
    """Copia mensagem para WhatsApp"""
    try:
        import pyperclip
        
        referencia = data_referencia.get_date()
        periodos = obter_proximos_periodos(referencia, 10)
        
        mensagem = "Conforme meu calendário de guarda, estarei com minhas filhas nos dias:\n\n"
        
        for periodo in periodos:
            mensagem += f"📅 {periodo['inicio'].strftime('%d/%m/%Y')} até {periodo['fim'].strftime('%d/%m/%Y')}\n"
        
        pyperclip.copy(mensagem)
        
        ctk.CTkLabel(
            app,
            text="✅ Mensagem copiada para a área de transferência!",
            text_color="green",
            font=("Segoe UI", 10)
        ).pack(pady=5)
    
    except ImportError:
        ctk.CTkLabel(
            app,
            text="❌ Instale pyperclip: pip install pyperclip",
            text_color="red",
            font=("Segoe UI", 10)
        ).pack(pady=5)
    except Exception as e:
        ctk.CTkLabel(
            app,
            text=f"❌ Erro: {str(e)}",
            text_color="red",
            font=("Segoe UI", 10)
        ).pack(pady=5)


# ==========================================
# INTERFACE - HISTÓRICO
# ==========================================

def ver_historico():
    """Mostra histórico de consultas"""
    janela = ctk.CTkToplevel(app)
    janela.title("Histórico de Consultas")
    janela.geometry("500x400")
    
    ctk.CTkLabel(
        janela,
        text="📋 HISTÓRICO DE CONSULTAS",
        font=("Segoe UI", 18, "bold")
    ).pack(pady=15)
    
    if not dados.historico:
        ctk.CTkLabel(
            janela,
            text="Nenhuma consulta registrada",
            font=("Segoe UI", 12)
        ).pack(pady=20)
    else:
        texto = ctk.CTkTextbox(janela, width=450, height=300)
        texto.pack(padx=20, pady=10, fill="both", expand=True)
        
        for i, consulta in enumerate(reversed(dados.historico[-20:]), 1):
            timestamp = datetime.fromisoformat(consulta['timestamp']).strftime('%d/%m/%Y %H:%M')
            texto.insert("end", f"{i}. {consulta['data']} - {timestamp}\n")
        
        texto.configure(state="disabled")


# ==========================================
# JANELA PRINCIPAL
# ==========================================

app = ctk.CTk()
app.title("👧 Controle de Guarda das Filhas")
app.geometry("850x750")
app.resizable(False, False)

# Tema
frame_tema = ctk.CTkFrame(app)
frame_tema.pack(pady=10, fill="x", padx=20)

ctk.CTkLabel(frame_tema, text="Tema:", font=("Segoe UI", 11)).pack(side="left", padx=5)

def mudar_tema(tema):
    ctk.set_appearance_mode(tema)
    dados.config['tema'] = tema
    dados.salvar_config()

ctk.CTkButton(
    frame_tema,
    text="☀ Claro",
    command=lambda: mudar_tema("light"),
    width=80
).pack(side="left", padx=3)

ctk.CTkButton(
    frame_tema,
    text="🌙 Escuro",
    command=lambda: mudar_tema("dark"),
    width=80
).pack(side="left", padx=3)

# Título
titulo = ctk.CTkLabel(
    app,
    text="👧 CONTROLE DE GUARDA DAS FILHAS",
    font=("Segoe UI", 26, "bold")
)
titulo.pack(pady=15)

# Frame principal
frame = ctk.CTkFrame(app)
frame.pack(padx=30, pady=10, fill="both", expand=True)

# Data referência
label_ref = ctk.CTkLabel(
    frame,
    text="📅 Data em que você estava com suas filhas",
    font=("Segoe UI", 14, "bold")
)
label_ref.pack(pady=(15, 5))

data_referencia = DateEntry(
    frame,
    locale="pt_BR",
    date_pattern="dd/mm/yyyy",
    width=20,
    background="#1F6AA5",
    foreground="white",
    borderwidth=2
)
data_referencia.pack(pady=5)

# Data evento
label_evento = ctk.CTkLabel(
    frame,
    text="📅 Data do evento",
    font=("Segoe UI", 14, "bold")
)
label_evento.pack(pady=(15, 5))

data_evento = DateEntry(
    frame,
    locale="pt_BR",
    date_pattern="dd/mm/yyyy",
    width=20,
    background="#1F6AA5",
    foreground="white",
    borderwidth=2
)
data_evento.pack(pady=5)

# Frame com botões principais
frame_botoes = ctk.CTkFrame(frame)
frame_botoes.pack(pady=20, fill="x")

botao_dashboard = ctk.CTkButton(
    frame_botoes,
    text="📊 Dashboard",
    command=criar_dashboard,
    width=120,
    height=40,
    font=("Segoe UI", 12, "bold")
)
botao_dashboard.pack(side="left", padx=5)

botao_verificar = ctk.CTkButton(
    frame_botoes,
    text="🔎 Verificar Evento",
    command=lambda: criar_pesquisa_data(),
    width=150,
    height=40,
    font=("Segoe UI", 12, "bold")
)
botao_verificar.pack(side="left", padx=5)

botao_calendario = ctk.CTkButton(
    frame_botoes,
    text="📆 Calendário Visual",
    command=criar_calendario_visual,
    width=150,
    height=40,
    font=("Segoe UI", 12, "bold")
)
botao_calendario.pack(side="left", padx=5)

# Frame com botões secundários
frame_botoes2 = ctk.CTkFrame(frame)
frame_botoes2.pack(pady=10, fill="x")

botao_lista = ctk.CTkButton(
    frame_botoes2,
    text="📋 Períodos",
    command=lambda: listar_periodos(52),
    width=110,
    height=40,
    font=("Segoe UI", 11, "bold")
)
botao_lista.pack(side="left", padx=5)

botao_pdf = ctk.CTkButton(
    frame_botoes2,
    text="📄 Exportar PDF",
    command=exportar_pdf,
    width=130,
    height=40,
    font=("Segoe UI", 11, "bold")
)
botao_pdf.pack(side="left", padx=5)

botao_excel = ctk.CTkButton(
    frame_botoes2,
    text="📊 Exportar Excel",
    command=exportar_excel,
    width=140,
    height=40,
    font=("Segoe UI", 11, "bold")
)
botao_excel.pack(side="left", padx=5)

botao_whatsapp = ctk.CTkButton(
    frame_botoes2,
    text="💬 WhatsApp",
    command=copiar_whatsapp,
    width=120,
    height=40,
    font=("Segoe UI", 11, "bold")
)
botao_whatsapp.pack(side="left", padx=5)

botao_historico = ctk.CTkButton(
    frame_botoes2,
    text="📋 Histórico",
    command=ver_historico,
    width=110,
    height=40,
    font=("Segoe UI", 11, "bold")
)
botao_historico.pack(side="left", padx=5)

# Resultado
resultado = ctk.CTkLabel(
    frame,
    text="",
    height=90,
    corner_radius=12,
    font=("Segoe UI", 13, "bold"),
    wraplength=600
)
resultado.pack(pady=20, padx=20, fill="x")

# Rodar aplicação
if __name__ == "__main__":
    app.mainloop()
